from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


OUTPUT_FILE = Path(__file__).with_name("Excel_Project_Tracker.xlsx")
TABLE_STYLE = "TableStyleMedium2"
DATA_ROW_COUNT = 500


REFERENCE_LISTS = {
    "Priority": ["Low", "Medium", "High", "Urgent"],
    "Status": ["Not Started", "In Progress", "Waiting", "Completed", "On Hold", "Cancelled"],
    "Software/Tech": [
        "ArcGIS",
        "Excel",
        "IMPLAN",
        "Word",
        "PowerPoint",
        "Outlook",
        "SQL",
        "Power BI",
        "Access",
        "Adobe Acrobat",
    ],
    "Assigned From": ["Manager", "Director", "Client", "Coworker", "Self-Initiated"],
    "Owner": ["Isaac"],
    "Tracked Field": [
        "Assigned Date",
        "Last Update Date",
        "Primary Due Date",
        "Status",
        "Priority",
        "Owner",
        "Scope",
        "Summary Notes",
        "Next Action",
    ],
    "Role": ["Requestor", "Decision Maker", "Reviewer", "Client Contact", "Internal Contact"],
}


PROJECT_HEADERS = [
    "Project ID",
    "Project Name",
    "Assigned From",
    "Assigned From (Additional)",
    "Assigned Date",
    "Last Update Date",
    "Primary Due Date",
    "Priority",
    "Status",
    "Percent Complete",
    "Completed Date",
    "Owner",
    "Department/Client",
    "Required Software/Tech",
    "Summary Notes",
    "Next Action",
    "Days Until Due",
    "Overdue Flag",
]

MILESTONE_HEADERS = [
    "Milestone ID",
    "Project ID",
    "Milestone Name",
    "Assigned By",
    "Milestone Due Date",
    "Milestone Status",
    "Milestone Priority",
    "Milestone Notes",
    "Completed Date",
    "Days Until Due",
    "Overdue Flag",
]

UPDATE_HEADERS = [
    "Update ID",
    "Project ID",
    "Update Date",
    "Updated Field",
    "Old Value",
    "New Value",
    "Reason/Change Note",
    "Entered By",
]

CONTACT_HEADERS = [
    "Contact ID",
    "Project ID",
    "Contact Name",
    "Organization",
    "Role",
    "Email",
    "Phone",
    "Notes",
]


DATE_FILL_OVERDUE = PatternFill(fill_type="solid", fgColor="F8CBAD")
DATE_FILL_SOON = PatternFill(fill_type="solid", fgColor="FFF2CC")
PRIORITY_FILLS = {
    "Low": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "Medium": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "High": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "Urgent": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}
STATUS_COMPLETE_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
STATUS_HOLD_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
STATUS_CANCELLED_FILL = PatternFill(fill_type="solid", fgColor="EAD1DC")


def add_reference_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Reference_Lists")
    for column_index, (title, values) in enumerate(REFERENCE_LISTS.items(), start=1):
        header_cell = sheet.cell(row=1, column=column_index, value=title)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center")

        for row_index, value in enumerate(values, start=2):
            sheet.cell(row=row_index, column=column_index, value=value)

        max_row = len(values) + 1
        range_ref = f"'Reference_Lists'!${get_column_letter(column_index)}$2:${get_column_letter(column_index)}${max_row}"
        defined_name = DefinedName(f"{sanitize_name(title)}List", attr_text=range_ref)
        workbook.defined_names.add(defined_name)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(len(title) + 2, 18)

    sheet.freeze_panes = "A2"


def sanitize_name(title: str) -> str:
    return title.replace("/", "").replace(" ", "")


def setup_sheet(workbook: Workbook, title: str, headers: list[str], table_name: str, widths: dict[str, int]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    blank_row = [""] * len(headers)
    sheet.append(blank_row)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True

    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(headers))}2")
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    note_columns = {"O", "P", "H", "G"}
    for row in sheet.iter_rows(min_row=2, max_row=DATA_ROW_COUNT, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column_letter in note_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_list_validation(sheet, cell_range: str, defined_name: str, prompt: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=f"={defined_name}",
        allow_blank=True,
        showDropDown=False,
        promptTitle="Select a value",
        prompt=prompt,
    )
    validation.add(cell_range)
    sheet.add_data_validation(validation)


def add_text_guidance(sheet, cell_range: str, prompt: str) -> None:
    validation = DataValidation(
        type="custom",
        formula1="TRUE",
        allow_blank=True,
        showInputMessage=True,
        promptTitle="Entry tip",
        prompt=prompt,
    )
    validation.add(cell_range)
    sheet.add_data_validation(validation)


def apply_project_formulas(sheet) -> None:
    sheet["Q2"] = '=IF(OR(G2="",I2="Completed"),"",G2-TODAY())'
    sheet["R2"] = '=IF(AND(G2<>"",I2<>"Completed",G2<TODAY()),"Overdue","")'


def apply_milestone_formulas(sheet) -> None:
    sheet["J2"] = '=IF(OR(E2="",F2="Completed"),"",E2-TODAY())'
    sheet["K2"] = '=IF(AND(E2<>"",F2<>"Completed",E2<TODAY()),"Overdue","")'


def apply_date_formats(sheet, date_columns: list[str]) -> None:
    for column in date_columns:
        for row in range(2, DATA_ROW_COUNT + 1):
            sheet[f"{column}{row}"].number_format = "m/d/yyyy"


def apply_percent_format(sheet, cell_range: str) -> None:
    for row in range(2, DATA_ROW_COUNT + 1):
        sheet[f"{cell_range[0]}{row}"].number_format = "0%"


def add_conditional_formatting(sheet, due_column: str, status_column: str, priority_column: str, end_row: int) -> None:
    due_range = f"{due_column}2:{due_column}{end_row}"
    sheet.conditional_formatting.add(
        due_range,
        FormulaRule(
            formula=[f'AND({due_column}2<>"",{status_column}2<>"Completed",{due_column}2<TODAY())'],
            stopIfTrue=True,
            fill=DATE_FILL_OVERDUE,
        ),
    )
    sheet.conditional_formatting.add(
        due_range,
        FormulaRule(
            formula=[f'AND({due_column}2<>"",{status_column}2<>"Completed",{due_column}2>=TODAY(),{due_column}2<=TODAY()+7)'],
            stopIfTrue=True,
            fill=DATE_FILL_SOON,
        ),
    )

    priority_range = f"{priority_column}2:{priority_column}{end_row}"
    for priority, fill in PRIORITY_FILLS.items():
        sheet.conditional_formatting.add(
            priority_range,
            CellIsRule(operator="equal", formula=[f'"{priority}"'], fill=fill),
        )

    status_range = f"{status_column}2:{status_column}{end_row}"
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Completed"'], fill=STATUS_COMPLETE_FILL),
    )
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"On Hold"'], fill=STATUS_HOLD_FILL),
    )
    sheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Cancelled"'], fill=STATUS_CANCELLED_FILL),
    )


def build_workbook() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_reference_sheet(workbook)

    setup_sheet(
        workbook,
        "Projects",
        PROJECT_HEADERS,
        "ProjectsTable",
        {
            "A": 14,
            "B": 28,
            "C": 18,
            "D": 24,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 12,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 16,
            "M": 20,
            "N": 26,
            "O": 36,
            "P": 28,
            "Q": 14,
            "R": 14,
        },
    )
    setup_sheet(
        workbook,
        "Milestones",
        MILESTONE_HEADERS,
        "MilestonesTable",
        {
            "A": 14,
            "B": 14,
            "C": 28,
            "D": 18,
            "E": 14,
            "F": 16,
            "G": 16,
            "H": 34,
            "I": 14,
            "J": 14,
            "K": 14,
        },
    )
    setup_sheet(
        workbook,
        "Updates_Log",
        UPDATE_HEADERS,
        "UpdatesLogTable",
        {
            "A": 14,
            "B": 14,
            "C": 14,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 42,
            "H": 18,
        },
    )
    setup_sheet(
        workbook,
        "Contacts",
        CONTACT_HEADERS,
        "ContactsTable",
        {
            "A": 14,
            "B": 14,
            "C": 24,
            "D": 22,
            "E": 18,
            "F": 28,
            "G": 18,
            "H": 32,
        },
    )

    projects = workbook["Projects"]
    milestones = workbook["Milestones"]
    updates = workbook["Updates_Log"]
    contacts = workbook["Contacts"]

    add_list_validation(projects, f"C2:C{DATA_ROW_COUNT}", "AssignedFromList", "Pick the main requestor for this project.")
    add_list_validation(projects, f"H2:H{DATA_ROW_COUNT}", "PriorityList", "Select the project's priority level.")
    add_list_validation(projects, f"I2:I{DATA_ROW_COUNT}", "StatusList", "Select the current project status.")
    add_list_validation(projects, f"L2:L{DATA_ROW_COUNT}", "OwnerList", "Choose the person responsible for the project.")
    add_text_guidance(projects, f"A2:A{DATA_ROW_COUNT}", "Use a unique project ID such as PRJ-001.")
    add_text_guidance(
        projects,
        f"D2:D{DATA_ROW_COUNT}",
        "Use this field for additional requestors, separated by commas if needed.",
    )
    add_text_guidance(
        projects,
        f"N2:N{DATA_ROW_COUNT}",
        "Enter one or more tools separated by commas, using names from Reference_Lists when possible.",
    )

    add_list_validation(milestones, f"D2:D{DATA_ROW_COUNT}", "AssignedFromList", "Pick the person who assigned the milestone.")
    add_list_validation(milestones, f"F2:F{DATA_ROW_COUNT}", "StatusList", "Select the milestone status.")
    add_list_validation(milestones, f"G2:G{DATA_ROW_COUNT}", "PriorityList", "Select the milestone priority.")
    add_text_guidance(milestones, f"A2:A{DATA_ROW_COUNT}", "Use a unique milestone ID such as M-001.")
    add_text_guidance(milestones, f"B2:B{DATA_ROW_COUNT}", "Enter the matching Project ID from the Projects sheet.")

    add_list_validation(updates, f"D2:D{DATA_ROW_COUNT}", "TrackedFieldList", "Select the field that changed.")
    add_text_guidance(updates, f"B2:B{DATA_ROW_COUNT}", "Enter the matching Project ID from the Projects sheet.")
    add_list_validation(contacts, f"E2:E{DATA_ROW_COUNT}", "RoleList", "Select this contact's role for the project.")
    add_text_guidance(contacts, f"B2:B{DATA_ROW_COUNT}", "Enter the matching Project ID from the Projects sheet.")

    apply_project_formulas(projects)
    apply_milestone_formulas(milestones)

    apply_date_formats(projects, ["E", "F", "G", "K"])
    apply_date_formats(milestones, ["E", "I"])
    apply_date_formats(updates, ["C"])
    apply_percent_format(projects, "J")

    add_conditional_formatting(projects, "G", "I", "H", DATA_ROW_COUNT)
    add_conditional_formatting(milestones, "E", "F", "G", DATA_ROW_COUNT)

    workbook.save(OUTPUT_FILE)


if __name__ == "__main__":
    build_workbook()
